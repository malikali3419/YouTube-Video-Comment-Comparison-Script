import pandas as pd
import requests
import re
import logging
import os, time
import openpyxl
from openpyxl.utils import get_column_letter

from datetime import datetime
from selenium import webdriver
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import uuid

logging.basicConfig(
    level=logging.DEBUG,  # Set the log level
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("script.log"),
        logging.StreamHandler()
    ]
)

# Suppress DEBUG logs from external libraries
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("selenium").setLevel(logging.WARNING)
logging.getLogger("requests").setLevel(logging.WARNING)

logger = logging.getLogger(__name__)


class GetUserInput:
    def __init__(self):
        self.excel_file_name = input("Please enter name of the Excel file (e.g., 'input_file' or 'input_file.xlsx'): ")
        if not self.excel_file_name.endswith('.xlsx'):
            self.excel_file_name += '.xlsx'
        file_path = os.path.join(os.getcwd(), self.excel_file_name)
        return file_path


class SeleniumFunctionality:

    def __init__(self):
        self.options = Options()
        self.options.add_argument('--disable-dev-shm-usage')
        self.options.add_argument("--headless")
        self.options.add_argument('--no-sandbox')
        self.options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"
        )
        self.chromedriver_path = ChromeDriverManager().install()

    def close(self):
        self.driver.quit()

    def process_comments_in_hadzy(self, video_link):
        try:
            self.driver = webdriver.Chrome(service=Service(self.chromedriver_path), options=self.options)
            self.driver.get(self.hadzy_url)
            try:
                accept_all_button = WebDriverWait(self.driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, "//span[text()='Accept all']"))
                )
                accept_all_button.click()
            except TimeoutException:
                logger.info("Accept cookies button not found so the scrapper is proceeding ....")

            input_div = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "MuiInputBase-root"))
            )
            input_link = WebDriverWait(input_div, 10).until(
                EC.visibility_of_element_located((By.TAG_NAME, 'input'))
            )
            input_link.send_keys(video_link)

            search_btn = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.CLASS_NAME, 'MuiButtonBase-root'))
            )
            search_btn.click()

            time.sleep(2)

            card_content_div = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "MuiCardContent-root"))
            )
            try:
                load_data_btn = WebDriverWait(card_content_div[0], 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='load data']"))
                )
                load_data_btn.click()
            except TimeoutException:
                logger.info("Load Data button not found so the scrapper is proceeding ....")

            comments_btn = WebDriverWait(self.driver, 50).until(
                EC.presence_of_element_located((By.XPATH, "//button[@aria-label='view comments']"))
            )

            comments_btn.click()

            time.sleep(2)
            self.close()
            if comments_btn:
                return
        except Exception as e:
            logger.error("Failed to load comments on hadzy....", e)
            return


class YouTubeCommentProcessor(GetUserInput, SeleniumFunctionality):

    def __init__(self, reprocess=False, reprocess_try=1):
        """
        Initializes the YouTubeCommentProcessor with the given file path.

        Args:
            file_path (str): The path to the Excel file containing video data.
        """
        if not reprocess:
            file_path = GetUserInput.__init__(self)
        else:
            self.excel_file_name = 'requested_links.xlsx'
            file_path = os.path.join(os.getcwd(), self.excel_file_name)

        self.file_path = file_path
        self.failed_video_found = False
        self.reprocess_try = reprocess_try
        try:
            self.data_frame = pd.read_excel(self.file_path)
            logging.info("File Loaded Successfully!")
            logging.info(self.data_frame.head())
            logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(message)s')

        except FileNotFoundError:
            logger.error(f"Error: The file '{self.file_path}' was not found.")
            raise
        except Exception as e:
            logger.error(f"Error: The file '{self.file_path}' was not found.")
            raise
        self.hadzy_url = "https://www.hadzy.com/"
        SeleniumFunctionality.__init__(self)

    @staticmethod
    def extract_video_id(url):
        """
        Extracts the video ID from a YouTube URL.

        Args:
            url (str): The YouTube URL.

        Returns:
            str: The extracted video ID or None if not found.
        """
        video_id = re.findall(r'(?:v=|\/)([0-9A-Za-z_-]{11}).*', url)
        return video_id[0] if video_id else None

    @staticmethod
    def get_video_statistics(video_id):
        """
        Fetches video statistics from the Hadzy API.

        Args:
            video_id (str): The YouTube video ID.

        Returns:
            dict: The response JSON containing video statistics.
        """

        url = f"https://www.hadzy.com/api/videos/{video_id}"
        response = requests.get(url)
        return response.json()

    @staticmethod
    def get_comments(video_id, video_link, page=1, size=50, show_logs=True):
        """
        Fetches comments from the Hadzy API for a given video.

        Args:
            video_id (str): The YouTube video ID.
            page (int): The page number to retrieve.
            size (int): The number of comments per page (default is 100).

        Returns:
            dict: The response JSON containing comments.
            :param show_logs:
        """
        url = (
            f"https://www.hadzy.com/api/comments/{video_id}?"
            f"page={page}&size={size}&sortBy=publishedAt&"
            f"direction=asc%20%20%20%20%20%20&searchTerms=&author="
            )
        response = requests.get(url)
        if show_logs:
            logger.info(f"Getting Comments from this: {video_link} on page number: {page}")
        return response.json()

    def process_video(self, video_id, video_link, title, upload_time):
        """
        Processes comments for a single video, fetching all pages up to 1000.

        Args:
            video_id (str): The YouTube video ID.

        Returns:
            tuple: A tuple containing video metadata and a list of comments.
            :param upload_time:
            :param title:
            :param video_id:
            :param video_link:
        """
        total_page_count = 0
        max_tries = 1
        while max_tries <= 2 and total_page_count == 0:
            self.process_comments_in_hadzy(video_link)
            comments_data = self.get_comments(
                video_id=video_id,
                page=0,
                size=50,
                video_link=video_link,
                show_logs=False
            )
            page_info = comments_data.get("pageInfo", "")
            if page_info:
                total_page_count = page_info.get("totalPages", 0)
            max_tries += 1

        comments_list = []
        if total_page_count == 0:
            return

        for page in range(0, min(total_page_count, 1000) + 1):
            comments_data = self.get_comments(
                video_id=video_id,
                page=page,
                size=50,
                video_link=video_link
            )
            comments_list.extend(comments_data['content'])

        upload_date, upload_time = self.split_datetime(upload_time)

        return title, upload_date, upload_time, comments_list

    @staticmethod
    def split_datetime(datetime_str):
        """
        Splits a datetime string into date and time components.

        Args:
            datetime_str (str): The datetime string in ISO 8601 format.

        Returns:
            tuple: A tuple containing the date and time as strings.
        """
        dt = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
        return dt.date().isoformat(), dt.time().isoformat()

    @staticmethod
    def delete_already_exits_file(self, file_name):
        file_path = os.path.join(os.getcwd(), file_name)
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"File {file_path} has been removed successfully.")
        else:
            print(f"File {file_path} does not exist.")

    def save_requested_video_row_to_excel(self, row):
        """
        Saves the row data to a new Excel file with specified headers.

        Args:
            row (pd.Series): A row from the DataFrame.
        """
        if self.reprocess_try == 2:
            file_name = f"requested_links_{2}.xlsx"
        else:
            file_name = f"requested_links.xlsx"
        headers = [
            'Channel_Name_A',
            'Video_Title_A',
            'Video_URL_A',
            'Channel_Name_B',
            'Video_Title_B',
            'Video_URL_B',
            'Similarity'
        ]

        # Check if file exists
        try:
            pd.read_excel(file_name)
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                row_to_add = {header: row[header] for header in headers}
                df_new = pd.DataFrame([row_to_add])
                df_new.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        except FileNotFoundError:
            # Create a new file
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                row_to_add = {header: row[header] for header in headers}
                df_new = pd.DataFrame([row_to_add])
                df_new.to_excel(writer, index=False, header=headers)

        self.failed_video_found = True

    @staticmethod
    def _get_comment_count(stats):
        """
        Extracts the comment count from video statistics.

        Args:
            stats (dict): Video statistics dictionary.

        Returns:
            int: The number of comments.
        """

        if stats and stats['items'][0]['statistics']['commentCount']:
            return int(stats['items'][0]['statistics'].get('commentCount', 0))
        return 0

    def _save_comments_to_excel(self, result_a, result_b, channel_a, index):
        """
        Saves the comments data to an Excel file.

        Args:
            result_a (tuple): The result of processing video A.
            result_b (tuple): The result of processing video B.
            channel_a (str): The channel name.
            index (int): The index of the row in the DataFrame.
        """
        title_a, upload_date_a, upload_time_a, comments_list_a = result_a
        title_b, upload_date_b, upload_time_b, comments_list_b = result_b

        upload_time_date_a = f"{upload_date_a} {upload_time_a}"
        upload_time_date_b = f"{upload_date_b} {upload_time_b}"
        max_comments_count = max(len(comments_list_a), len(comments_list_b))

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Comments'

        headers = [
            'Video_Title_A', 'Video_Upload_DateTime_A', 'Video_Comment_A', 'Comment_Date_A', 'Comment_Time_A',
            'Comment_A_Author_Name', 'Comment_A_Like_Count',
            'Video_Title_B', 'Video_Upload_DateTime_B', 'Video_Comment_B', 'Comment_Date_B', 'Comment_Time_B',
            'Comment_B_Author_Name', 'Comment_B_Like_Count'
        ]

        for col_num, header in enumerate(headers, 1):
            sheet[f'{get_column_letter(col_num)}1'] = header

        for i in range(max_comments_count):
            row_data = self._get_row_data(
                i,
                comments_list_a,
                comments_list_b,
                title_a,
                upload_time_date_a,
                title_b,
                upload_time_date_b
            )
            for col_num, value in enumerate(row_data, 1):
                sheet[f'{get_column_letter(col_num)}{i + 2}'] = value

        file_name = f'{channel_a.strip().replace(" ", "_")}_{index + 1}_{max_comments_count}.xlsx'
        workbook.save(file_name)

    def _get_row_data(
            self,
            index,
            comments_list_a,
            comments_list_b,
            title_a,
            upload_time_date_a,
            title_b,
            upload_time_date_b
    ):
        """
        Retrieves row data for the Excel file.

        Args:
            index (int): The index of the comment in the list.
            comments_list_a (list): List of comments for video A.
            comments_list_b (list): List of comments for video B.
            title_a (str): Title of video A.
            upload_time_date_a (str): Upload datetime of video A.
            title_b (str): Title of video B.
            upload_time_date_b (str): Upload datetime of video B.

        Returns:
            list: The row data to be written to the Excel file.
        """
        row_data = []

        if index < len(comments_list_a):
            comment_a = comments_list_a[index]
            comment_date_a, comment_time_a = self.split_datetime(comment_a['publishedAt'])
            row_data.extend([
                title_a, upload_time_date_a,
                comment_a['textDisplay'], comment_date_a, comment_time_a,
                comment_a['authorDisplayName'], comment_a['likeCount']
            ])
        else:
            row_data.extend([''] * 7)

        if index < len(comments_list_b):
            comment_b = comments_list_b[index]
            comment_date_b, comment_time_b = self.split_datetime(comment_b['publishedAt'])
            row_data.extend([
                title_b, upload_time_date_b,
                comment_b['textDisplay'], comment_date_b, comment_time_b,
                comment_b['authorDisplayName'], comment_b['likeCount']
            ])
        else:
            row_data.extend([''] * 7)

        return row_data

    def process_row(self, row, index):
        """
        Processes a single row of the Excel file, checks video statistics, and saves comments if conditions are met.

        Args:
            row (pd.Series): A row from the DataFrame.
            index (int): The index of the row in the DataFrame.
        """
        video_id_a = self.extract_video_id(row['Video_URL_A'])
        video_id_b = self.extract_video_id(row['Video_URL_B'])
        channel_a = row['Channel_Name_A']

        if not video_id_a or not video_id_b:
            logger.warning(f"Missing video ID for row {index + 1}: Video A ID={video_id_a}, Video B ID={video_id_b}")
            return

        stats_a = self.get_video_statistics(video_id_a)
        stats_b = self.get_video_statistics(video_id_b)
        if stats_a is not None or stats_b is not None:
            title_a = stats_a['items'][0]['snippet']['title']
            upload_time_a = stats_b['items'][0]['snippet']['publishedAt']
            title_b = stats_b['items'][0]['snippet']['title']
            upload_time_b = stats_b['items'][0]['snippet']['publishedAt']

        comment_count_a = self._get_comment_count(stats_a)
        comment_count_b = self._get_comment_count(stats_b)

        if comment_count_a >= 50 and comment_count_b >= 50:

            result_a = self.process_video(video_id_b, row['Video_URL_A'], title_a, upload_time_a)
            result_b = self.process_video(video_id_b, row['Video_URL_B'], title_b, upload_time_b)

            if result_a is None or result_b is None:
                self.save_requested_video_row_to_excel(row)
                return

            self._save_comments_to_excel(result_a, result_b, channel_a, index)
        else:
            logger.info(
                f"In row {index + 1}, one or both videos have less than 50 comments: A={comment_count_a}, B={comment_count_b}")

    def process_file(self):
        """
        Processes the entire Excel file, row by row.
        """
        for index, row in self.data_frame.iterrows():
            self.process_row(row, index)
        if self.failed_video_found and self.reprocess_try != 2:
            reprocess = input("Do you want to reprocess the failed videos (Y/N)")
            if reprocess.lower() == "y":
                try:
                    processor = YouTubeCommentProcessor(True, reprocess_try=2)
                    processor.process_file()
                except Exception as e:
                    logger.error(f"An error occurred: {e}")
            else:
                return
        else:
            return


if __name__ == "__main__":
    try:
        processor = YouTubeCommentProcessor()
        processor.process_file()
    except Exception as e:
        logger.error(f"An error occurred: {e}")
